import openpyxl
import os
from pathlib import Path
from contextlib import contextmanager
from datetime import datetime
from openpyxl.styles import Font, Alignment

from .models import ProcessingRequest, ProcessingResult, Columns


class ExcelProcessor:
    def process(self, request: ProcessingRequest) -> ProcessingResult:
        try:
            validation_result = self._validate_request(request)
            if not validation_result.success:
                return validation_result

            request.source_path = self._normalize_path(request.source_path)
            request.target_path = self._normalize_path(request.target_path)

            with self._load_workbook_context(request.source_path) as (wb, sheet):
                required_columns = [col.value for col in Columns]
                header_info = self._get_index_filter_columns(sheet, request, required_columns)

                filter_result = self._filter_data(sheet, request, header_info, required_columns)

                if isinstance(filter_result, ProcessingResult) and filter_result.success is False:
                    return filter_result

            return ProcessingResult(
                success=True,
                message="Документ обработан",
                output_path=request.target_path
            )

        except Exception as e:
            return ProcessingResult(success=False, message=str(e), output_path=None)

    def _validate_request(self, request: ProcessingRequest) -> ProcessingResult:
        if not request.source_path:
            return ProcessingResult(
                success=False,
                message="Не указан путь к исходному файлу",
                output_path=None
            )

        source_path = Path(request.source_path)

        if not source_path.exists():
            return ProcessingResult(
                success=False,
                message=f"Файл не найден: {request.source_path}",
                output_path=None
            )

        if not source_path.is_file():
            return ProcessingResult(
                success=False,
                message=f"Указанный путь не является файлом: {request.source_path}",
                output_path=None
            )

        if not request.target_path:
            return ProcessingResult(
                success=False,
                message="Не указан путь для сохранениия результата",
                output_path=None
            )

        target_path = Path(request.target_path)

        if not request.source_path.lower().endswith('xlsx'):
            return ProcessingResult(
                success=False,
                message="Поддерживаются только файлы Excel ('*.xlsx')",
                output_path=None
            )

        if not os.access(request.source_path, os.R_OK):
            return ProcessingResult(
                success=False,
                message=f"Нет прав на чтение файла: {request.source_path}",
                output_path=None
            )

        if not request.filter_column:
            return ProcessingResult(
                success=False,
                message="Не указано столбец для фильтрации",
                output_path=None
            )

        if not request.filter_item:
            return ProcessingResult(
                success=False,
                message="Не указано значение для фильтрации",
                output_path=None
            )

        target_parent = target_path.parent
        if not target_path.exists():
            try:
                target_parent.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                return ProcessingResult(
                    success=False,
                    message=f"Не удалось создать целевую директорию: {str(e)}",
                    output_path=None
                )

        if not os.access(str(target_parent), os.W_OK):
            return ProcessingResult(
                success=False,
                message=f"Нет прав на записб в директорию: {target_parent}",
                output_path=None
            )

        return ProcessingResult(success=True, message="Валидация пройдена", output_path=None)

    def _filter_data(self, sheet, request, header_info, required_columns):
        try:
            with self._save_file_context(request.target_path) as (wb_out, sheet_out, start_row):
                self._format_header(sheet_out, start_row, required_columns)
                min_row_table = header_info[1] + 1
                actual_filter_column_lower = {key.lower(): value for key, value in header_info[0].items()}
                index_filter_column = actual_filter_column_lower[request.filter_column]

                date_start_row = start_row + 1
                found = False

                for row_value in sheet.iter_rows(values_only=True, min_row=min_row_table):
                    value = row_value[index_filter_column]
                    if self._compare_values(value, request.filter_item):
                        found = True
                        filtered_row = [row_value[header_info[0][col]] for col in required_columns]
                        for col_idx, cell_item in enumerate(filtered_row, start=1):
                            sheet_out.cell(row=date_start_row, column=col_idx, value=cell_item)
                        date_start_row += 1

                if not found:
                    wb_out.close()
                    if os.path.exists(request.target_path):
                        os.remove(request.target_path)
                    return ProcessingResult(
                        success=False,
                        message=f"Нет совпадений для фильтрации по значению '{request.filter_item}' в колонке '{request.filter_column}'", # noqa 501
                        output_path=None
                    )
                return ProcessingResult(
                    success=True,
                    message="Найдены совпадения",
                    output_path=request.target_path
                )

        except Exception as e:
            return ProcessingResult(
                success=False,
                message=f"Ошибка при фильтрации данных {str(e)}",
                output_path=None
            )

    def _compare_values(self, cell_value, filter_item):
        if cell_value is None and filter_item is None:
            return True

        if cell_value is None or filter_item is None:
            return False
        cell_type = type(cell_value)
        filter_type = type(filter_item)

        if cell_type == filter_type:
            if isinstance(cell_value, str):
                return cell_value.strip().lower() == filter_item.strip().lower()
            else:
                return cell_value == filter_item

        try:
            cell_folat = float(cell_value)
            filter_float = float(filter_item)
            return cell_folat == filter_float
        except (ValueError, TypeError):
            pass

        try:
            cell_str = str(cell_value).strip().lower()
            filter_str = str(filter_item).strip().lower()
            return cell_str == filter_str
        except (ValueError, TypeError):
            pass

        return False

    def _get_index_filter_columns(self, sheet, request, required_columns):
        """
        Находим строку с заголовками в которой содержится название столбца из filter_column.
        Возвращаем кортеж: (columns_index, row_index)
        """
        for row_index, row_value in enumerate(sheet.iter_rows(values_only=True), 1):
            if not row_value:
                continue
            row_columns_set = {cell for cell in row_value if cell is not None}
            required_columns_set = set(required_columns)
            # Ищем строку содержащюю столбец фильтрации
            if required_columns_set.issubset(row_columns_set):
                columns_index = {}
                # Собираем инедксы всех требуемых столбцов
                for column_name in required_columns:
                    if column_name in row_value:
                        columns_index[column_name] = row_value.index(column_name)
                return columns_index, row_index

        raise ValueError(f"Столбец '{request.filter_column}' не найден в файле или отсутвуют обязательные столбцы")

    @contextmanager
    def _load_workbook_context(self, source_path: str):
        wb = None
        try:
            file_size = os.path.getsize(source_path)
            max_size = 50 * 1024 * 1024  # 50MB

            if file_size > max_size:
                raise ValueError(f"Файл слишком болшой {file_size / (1024 * 1024):.1f}MB (максимум 50MB)")

            if file_size == 0:
                raise ValueError("Файл пустой")

            wb = openpyxl.load_workbook(source_path, read_only=True)
            sheet = wb.active
            if not sheet:
                raise ValueError("В файле нет активного листа")

            yield wb, sheet

        except Exception as e:
            raise Exception(f"Ошибка при загрузке файла: {str(e)}")

        finally:
            if wb:
                wb.close()

    @contextmanager
    def _save_file_context(self, target_path: str):
        wb_out = None
        try:
            os.makedirs(os.path.dirname(target_path), exist_ok=True)
            wb_out = openpyxl.Workbook()
            sheet_out = wb_out.active
            sheet_out.title = "Отфильтрованные данные"

            sheet_out['A1'] = 'Тип файла:'
            sheet_out['A1'].font = Font(bold=True)
            sheet_out['B1'] = 'Excel файл'
            sheet_out['A2'] = 'Дата формирования'
            sheet_out['A2'].font = Font(bold=True)
            sheet_out['B2'] = datetime.now().strftime('%d.%m.%Y %H:%M')

            start_row = 4
            yield wb_out, sheet_out, start_row

            wb_out.save(target_path)

        except Exception as e:
            raise Exception(f"Ошибка при сохранении файла: {str(e)}")

        finally:
            if wb_out:
                wb_out.close()

    def _normalize_path(self, path: str):
        if not path:
            return path

        path_obj = Path(path)

        if not path_obj.is_absolute():
            path_obj = Path.cwd() / path_obj

        return str(path_obj.resolve())

    def _format_header(self, sheet, header_row_idx, headers):
        bold_font = Font(bold=True)
        alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=header_row_idx, column=col_idx, value=header)
            cell.font = bold_font
            cell.alignment = alignment

        for col_idx, header in enumerate(headers, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = len(header) + 5
